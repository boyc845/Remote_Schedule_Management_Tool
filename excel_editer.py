from openpyxl import load_workbook
import datetime
import file_console

def str_to_date_format(date: str) -> 'date':
    '''convert str format to datetime format for date'''
    return datetime.date(year = int(date.split('/')[0]),
                         month = int(date.split('/')[1]),
                         day = int(date.split('/')[2]))

def search_specific_word(word: str, sheet: 'sheet') -> '(row, col)':
    '''search all cells which contain the word in a sheet and then return row and col'''
    for row in range(1,sheet.max_row+1):
        for col in range(1,sheet.max_column+1):
            if sheet.cell(row = row, column = col).value != None:
                if word in str(sheet.cell(row = row, column = col).value):
                    return(row, col)
    return None
    

def search_specific_word_as_list(word: str, sheet: 'sheet') -> list:
    '''search all cells which contain the word in a sheet and then return the list of
    row and col of the word's cell'''
    list_related_cell = []
    for row in range(1, sheet.max_row+1):
        for col in range(1,sheet.max_column+1):
            if sheet.cell(row = row, column = col).value != None:
                if word in str(sheet.cell(row = row, column = col).value):
                    list_related_cell.append((row, col))
    return list_related_cell

def write_date_to_excel(sheet: 'sheet', dict_data: dict, row: int):
    '''write data to sheet'''
    delivery_date = sheet.cell(row = row, column = search_specific_word('货期', sheet)[1]).value
    for key in dict_data.keys():
        if search_specific_word(key, sheet) != None:
            sheet.cell(row = row, column = search_specific_word(key, sheet)[1],
                            value = (delivery_date.date() - datetime.timedelta(days=int(dict_data[key]))))

def find_ready_to_write_row(sheet: 'sheet') -> list:
    '''find rows which need to be fill with date'''
    col_delivery_date = search_specific_word('货期', sheet)[1]
    col_chuban_date = search_specific_word('日期', sheet)[1]
    list_rows = []
    for row in range(1, sheet.max_row+1):
        if (sheet.cell(row = row, column = col_delivery_date).value != None) and\
           (sheet.cell(row = row, column = col_chuban_date).value == None):
            list_rows.append(row)
    return list_rows
    
def choose_template(style: str) -> dict:
    '''choose template'''
    try:
        dict_data = file_console.read_data_as_dict('模板_{}.txt'.format(style))
    except:
        dict_data = file_console.read_data_as_dict('模板_文胸.txt')
    return dict_data

def search_empty_row(sheet: 'sheet') -> int:
    '''search an empty row and return row number'''
    for index in range(1, sheet.max_row+1):
        for cell in sheet.rows[index]:
            if cell.value != None:
                break
            elif cell == sheet.rows[index][-1] and cell.value == None:
                return index+1

def backup_excelfile(excel_name: str, excel_object: 'excel'):
    '''back up excel file with name in backup directory'''
    path = file_console.create_folder_current_directory('excel_backup')
    p = path/excel_name
    excel_object.save(str(p))
    
class Excel_editer:
    def __init__(self, excel_name: str, sheetnum = 0):
        self.excel_name = excel_name
        self.excel = load_workbook(filename = excel_name)
        self.sheet = self.excel.worksheets[sheetnum]
        backup_excelfile(excel_name, self.excel)

    def write_date_to_rows(self):
        '''fill data to sheet'''
        for row in find_ready_to_write_row(self.sheet):
            dict_data = choose_template(self.sheet.cell(row = row,
                                                   column = search_specific_word('类型', self.sheet)[1]).value)
            write_date_to_excel(self.sheet, dict_data, row)
##########
    def write_done(self, dict_commands, receive_date):
        '''write done in specific cell'''
        for key in dict_commands.keys():
            for command in dict_commands[key]:
                if search_specific_word(command, self.sheet) == None or \
                   search_specific_word(key, self.sheet) == None:
                    return False
                col = search_specific_word(command, self.sheet)[1]
                row = search_specific_word(key, self.sheet)[0]
                print(row, col)
                self.sheet.cell(row = row, column = col, value = str(receive_date)+ '完成')
                

    def save_excelfile(self):
        '''save object to excel file'''
        self.excel.save(self.excel_name)

    def return_sheet(self):
        print(self.sheet)
        return self.sheet
    


if __name__ == '__main__':
    filename = '跟单进度跟踪 2016-3-24.xlsx'
    excel_editer = Excel_editer(filename)
    excel_editer.write_date_to_rows()
    excel_editer.save_excelfile()

##'跟单进度跟踪 2016-3-24.xlsx'
    msg = '完成 WILLOW 初板寄出 初板评语 主料确认\r\n\r\n2016-0'
    d = get_command_from_message(msg)
